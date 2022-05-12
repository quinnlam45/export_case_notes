from asyncore import write
from cgi import test
from fileinput import filename
import io
from json import load
from tempfile import NamedTemporaryFile
from turtle import mode

from numpy import byte
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook, save_workbook, ExcelWriter
from openpyxl.styles import Font
import pandas as pd
from excel_module import *

from pivot_table import PivotTable, make_pivot_table, pivot_df


dataframe_columns = ['CaseID', 'Referrer', 'Area', 'Risk']
dataframe_data = [(1, 'MASH', 'Sandwell', 'High'), (2, 'School', 'Dudley', 'Medium'), (3, 'MARAC', 'Sandwell','High')]
df = pd.DataFrame(dataframe_data, columns=dataframe_columns)


df_columns = {'High': [2, 5, 6], 'Medium': [7, 3, 2], 'Standard': [0, 3, 2]}
df_index = ['Sandwell', 'Dudley', 'Walsall']
df_dummy_pivot = pd.DataFrame(df_columns, index=df_index).rename_axis(index="Area", columns="Risk")



def return_no_of_df_rows(df):
    no_of_dimensions = df.ndim
    no_of_indexes = df.shape[0] # dataframe rows
    no_rows_in_df = no_of_dimensions + no_of_indexes

    return no_rows_in_df

def return_pivoted_dfs(df, df_rows, df_cols):
    dfs_list = []
    
    for df_row in df_rows:
        pivoted_df = pivot_df(df, df_row, df_cols)
        transformed_pivot_df = make_pivot_table(pivoted_df)
        dfs_list.append(transformed_pivot_df)
    
    return dfs_list


cols = 'Service risk level'
rows_to_pivot = ['Service area', 'Referrer', 'Gender', 'Ethnicity']

risk_category = ['High Risk', 'Medium Risk', 'Standard Risk']

filenames = []

for risk_level in risk_category:
    file_name = risk_level.replace(' ', '_')    
    file_name += '.xlsx'
    filenames.append(file_name)

ex_file = 'test_output.xlsx'



def write_to_excel_multiple_sheets(file_obj):
    with pd.ExcelWriter(file_obj) as writer:
        wb = writer.book

        for f in filenames: # run each query
            data = read_excel_file(f) # get data method
            df = pd.DataFrame(data[1:], columns=data[0])

            pivoted_df_list = return_pivoted_dfs(df, rows_to_pivot, cols)

            sh_name = f[:-5]

            current_row = 0

            for df in pivoted_df_list:
                no_of_cols_in_df = len(df.columns) + 1
                no_of_rows_in_df = len(df.index) + 1
                current_row += 1
                df.to_excel(writer, startrow=current_row, sheet_name=sh_name, na_rep=0)                
                
                ws = wb[sh_name]
                apply_table_styling(ws, current_row, no_of_rows_in_df, no_of_cols_in_df)
                
                current_row += no_of_rows_in_df
            # adjust col widths
            

        writer.save()


def build_report(excel_writer):

    current_row = 0
    buffer = io.BytesIO()

    with NamedTemporaryFile(suffix='.xlsx') as tmp:

        #implement different excelwriter/to excel methods here
        excel_writer(tmp)

        tmp.seek(0)
        stream = tmp.read()
        buffer.write(stream)

    return buffer

def bytes_to_file(stream, filename):
    with open(filename, 'wb') as ex_file:
        stream.seek(0)
        ex_file.write(stream.read())

def df_to_excel(df, dest_filename):

    with pd.ExcelWriter(dest_filename) as writer:
        df.to_excel(writer)
        writer.save()

stream = build_report(write_to_excel_multiple_sheets)
bytes_to_file(stream, ex_file)

def apply_ind_styles(f):
    wb = Workbook()
    ws = wb.active
    ws['A4'] = 'testtttttttttttt'
    ws['A5'] = 123
    ws['A6'] = 0.123
    ws['B4'] = 'testtttttttttttt'
    ws['B5'] = 123
    ws['B6'] = 0.123

    apply_column_width(ws, 5, 6, 2, ws_min_row=1)

    wb.save(f)

def try_with_df(f):
    df_to_excel(df_dummy_pivot, f)
    wb = load_workbook(f)
    ws = wb.active
    apply_table_styling(ws, 1, 4, 4)
    wb.save(f)

# apply_ind_styles(ex_file)
# try_with_df(ex_file)

def func_test(col_range=(1, 3)):
    col_min = col_range[0]
    col_max = col_range[1]
    print(col_min)
    print(col_max)

# func_test(col_range=(5, 7))