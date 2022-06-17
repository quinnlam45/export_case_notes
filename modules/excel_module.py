import io
import os
import string
import random
import re
from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill, Color, numbers
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.page import PrintPageSetup, PageMargins
from openpyxl.worksheet.header_footer import HeaderFooter, HeaderFooterItem
from openpyxl.writer.excel import ExcelWriter
import pandas as pd

# reads excel file and returns python data object/fake SQL data
def read_excel_file(filename):
    wb = load_workbook(filename)
    ws = wb.active

    data_rows = []

    for row in ws.iter_rows(min_row=1, max_row=1):
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        data_rows.append(row_list)

    for row in ws.iter_rows(min_row=2):
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        data_rows.append(tuple(row_list))
    return data_rows

def df_to_excel(df, sh_name, dest_filename):
    wb = load_workbook(dest_filename)
    ws = wb.active
    row = ws.max_row + 1

    with pd.ExcelWriter(dest_filename, mode='a', if_sheet_exists="overlay") as writer:
        df.to_excel(writer, startrow=row, sheet_name=sh_name)
        writer.save()

def df_to_IO(df):
    stream = io.BytesIO()
    with pd.ExcelWriter(stream) as writer:
        df.to_excel(writer)
    
    return stream

def remove_file_if_exists(filename):
    if os.path.exists(filename):
        os.remove(filename)

def read_file_return_bytes(filename):
    excel_file_bytes = io.BytesIO()

    if os.path.exists(filename):
        with open(filename, 'rb') as excel_file:
            excel_file.seek(0)
            bytes_stream = excel_file.read()
            excel_file_bytes.write(bytes_stream)
            excel_file.close()

        excel_file_bytes.seek(0)
        return excel_file_bytes
    
    else:
        raise Exception(f'File: {filename} does not exist')

def create_random_pw_string():
    no_of_characters = random.randint(3,5)
    no_of_letters = random.randint(5,6)
    
    characters = string.ascii_letters + string.digits + '!£$%^*_#~-.@'
    pw_letters = string.ascii_letters
    
    rand_char = ''.join(random.choice(characters) for i in range(no_of_characters))
    rand_letters = ''.join(random.choice(pw_letters) for i in range(no_of_letters))
    
    pw_string = rand_char + rand_letters
    
    # shuffle characters
    pw_list = list(pw_string)
    random.SystemRandom().shuffle(pw_list)
    pw_string = ''.join(pw_list)
    
    return pw_string

def create_temp_file_name(string_length=10):
    characters = string.ascii_letters + string.digits
    rand_char = ''.join(random.choice(characters) for i in range(string_length))
    return rand_char

def set_password_for_excel(filename, pw):
    from pathlib import Path
    import subprocess

    excel_file_path = Path(filename)

    vbs_script = \
    f"""' Save with password required upon opening

    Set excel_object = CreateObject("Excel.Application")
    Set workbook = excel_object.Workbooks.Open("{excel_file_path}")

    excel_object.DisplayAlerts = False
    excel_object.Visible = False

    workbook.SaveAs "{excel_file_path}",, "{pw}"

    workbook.Close()

    excel_object.Application.Quit
    """

    # write
    vbs_script_path = excel_file_path.parent.joinpath("set_pw.vbs")
    with open(vbs_script_path, "w") as file:
        file.write(vbs_script)

    #execute
    subprocess.call(['cscript.exe', str(vbs_script_path)])

    # remove
    vbs_script_path.unlink()

    return None

def create_case_notes_excel_file(data, clientID, filename, pw_str):
    sql_data = data
    remove_file_if_exists(filename)

    try:
        wb = Workbook()
        ws1 = wb.active

        # add client ref details
        ws1["A1"] = "Client ref"
        ws1["A1"].font = Font(bold=True)
        bd = Side(border_style="thin")
        ws1["A1"].border = Border(top=bd, bottom=bd)
        ws1["A2"] = clientID    
        ws1["A2"].fill = PatternFill("solid", fgColor=Color(indexed=22))
    
        max_col_width = 10
    
        # add data
        start_row = ws1.max_row + 1
        for row_no, datarow in enumerate(sql_data, start=1):
            for col_no, value in enumerate(datarow, start=1):
                ws1.cell(row=start_row+row_no, column=col_no, value=value)
                # update max col width
                if col_no == 1 and len(value) > max_col_width:
                    max_col_width = len(value)

        max_row = ws1.max_row
        start_row += 1
        max_col_letter = ws1.cell(column=ws1.max_column, row=1).column_letter
    
        # set width of last col
        ws1.column_dimensions[max_col_letter].width = 80
    
        # convert data range into table
        datarange = f"A{start_row}:{max_col_letter}{max_row}"
        tab = Table(displayName="Table1", ref=datarange)

        style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws1.add_table(tab)

        # adjust other col width
        ft = Font(b=True)
        for col in ws1.iter_cols(min_row=4, max_col=ws1.max_column-1, max_row=max_row):
            col_width = 9
            for cell in col:
                cell.alignment = Alignment(vertical="top")
                if isinstance(cell.value, str) and len(cell.value) > col_width:
                    col_width = len(cell.value)
            ws1.column_dimensions[cell.column_letter].width = col_width

        for row in ws1.iter_rows(min_row=start_row, min_col=ws1.max_column, max_col=ws1.max_column, max_row=max_row):
            for row_cell in row:
                row_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
        # page and print setup
        ws1.page_setup = PrintPageSetup(orientation = "landscape", scale=70, paperSize=9)
        ws1.page_margins = PageMargins(left=0.6, right=0.6, top=0.7, bottom=0.7)
        ws1.oddFooter.right.text = "&[Page]"
        ws1.evenFooter.left.text = "&[Page]"
    
        wb.save(filename)
        wb.close()

        set_password_for_excel(filename, pw_str)

        excel_file_bytes = read_file_return_bytes(filename)        

        return excel_file_bytes
    
    except:
        raise Exception('Unable to create file')
    finally:
        remove_file_if_exists(filename)


def reset_formatting_style(cell):
    ft = Font(bold=False)
    bd_style = Side(border_style=None)
    border_setting = Border(top=bd_style, left=bd_style, right=bd_style, bottom=bd_style)

    cell.font = ft
    cell.border = border_setting

def add_row_banding_style(cell):    
    cell.style = "20 % - Accent1"
    cell.font = Font(size=11)

def set_column_header_style(cell):
    cell.style = "Total"
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

def set_row_header_style(cell):
    cell.alignment = Alignment(horizontal='left', vertical="top")

def return_max_column_width(cell_value, min_col_width_setting):
    if len(cell_value) > min_col_width_setting:
        max_width = len(cell_value)
    else:
        max_width = min_col_width_setting
    
    return max_width

def apply_excel_percent_formatting(cell):
    cell_val = cell.value
    # remove % sign and convert
    cell.value = float(cell_val[:-1])
    cell.number_format = numbers.FORMAT_PERCENTAGE

def apply_column_width(ws, min_col_width_setting, ws_min_row, ws_max_row, ws_max_col, ws_min_col=1, header_styling=False, fixed_width=False):
    max_width = min_col_width_setting
    for column in ws.iter_cols(min_row=ws_min_row, max_row=ws_max_row, min_col=ws_min_col, max_col=ws_max_col):
        for column_cell in column:
            
            if header_styling == True:
                set_row_header_style(column_cell)
            
            if column_cell != None and isinstance(column_cell.value, str):
                if fixed_width == False:
                    max_width = return_max_column_width(column_cell.value, max_width)
                else:
                    max_width = min_col_width_setting
        
        ws.column_dimensions[column_cell.column_letter].width = max_width

def apply_table_styling(ws, start_row_no, max_row_no, max_col_no, min_col_no=1, first_col_min_width=25, other_col_min_width=10):
    if ws == None:
        raise Exception('No worksheet object passed')

    row_counter = 0
    # iter_rows is 1 based: +1 to min_row parameter
    for row in ws.iter_rows(min_row=start_row_no+1, max_row=start_row_no+max_row_no, min_col=min_col_no, max_col=max_col_no):
        for row_cell in row:
            reset_formatting_style(row_cell)
            
            if row_counter % 2:
                add_row_banding_style(row_cell)
            
            # if first row
            if row_counter == 0:
                set_column_header_style(row_cell)
            
            cell_val = row_cell.value
            if isinstance(cell_val, str) and re.search('\d%', cell_val) != None:
                apply_excel_percent_formatting(row_cell)

        row_counter += 1
    
    first_col_width = first_col_min_width
    remaining_col_width = other_col_min_width
    # adjust first column width
    apply_column_width(ws, first_col_width, start_row_no, start_row_no+max_row_no, ws_max_col=1, header_styling=True)
    # adjust remaining column width
    apply_column_width(ws, remaining_col_width, start_row_no, start_row_no+max_row_no, ws_min_col=min_col_no+1, ws_max_col=max_col_no, fixed_width=True)
